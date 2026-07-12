#!/usr/bin/python3

import os
import re
import sys
from tkinter import Tk, filedialog, messagebox

import openpyxl
import pandas as pd
import pdfplumber


def clean_money(value):
    return float(value.replace(",", ""))


def build_lines(page):
    words = page.extract_words(x_tolerance=2, y_tolerance=2, keep_blank_chars=False)
    rows = {}
    for word in words:
        y = round(word["top"])
        rows.setdefault(y, []).append(word)
    lines = []
    for y in sorted(rows):
        row = sorted(rows[y], key=lambda w: w["x0"])
        text = " ".join(word["text"] for word in row)
        lines.append(text)
    return lines


def parse_funds(lines):
    # parse the account funds
    rows = []
    account = 0  # 0 = main account
    for line in lines:
        tokens = line.split()
        # print(tokens)
        if len(tokens) < 4:
            continue
        first_token = tokens[0]
        # pull sub account id if present
        if first_token == "FUND" and len(tokens) == 8:
            account = tokens[7]
        # account name
        if tokens[3] == "DESCRIPTION:":
            account_name = (
                "Unit" if "Unit" in tokens else ",".join(tokens[4:]).replace("/", ",")
            )
        # we only want rows with object codes
        if not re.compile(r"^\d{4}$").match(first_token):
            continue
        numeric_tokens = [
            token for token in tokens if re.compile(r"^-?\d[\d,]*\.\d{2}$").match(token)
        ]
        if len(numeric_tokens) < 5:
            continue
        try:
            financial_plan = clean_money(numeric_tokens[-5])
            reconciled = clean_money(numeric_tokens[-4])
            unreconciled = clean_money(numeric_tokens[-3])
            total_obligations = clean_money(numeric_tokens[-2])
            balance_available = clean_money(numeric_tokens[-1])
        except ValueError:
            continue
        description_tokens = []
        for token in tokens[1:]:
            if token == numeric_tokens[-5]:
                break
            description_tokens.append(token)
        description = " ".join(description_tokens)
        rows.append(
            {
                "ObjectCode": int(first_token),
                "Description": description,
                "FinancialPlan": financial_plan,
                "Reconciled": reconciled,
                "Unreconciled": unreconciled,
                "TotalObligations": total_obligations,
                "BalanceAvailable": balance_available,
            }
        )
    return rows, account, account_name


def parse_comments(lines):
    # parse the comments
    rows = []
    plan = 0  # HACKY
    for line in lines:
        tokens = line.split()
        # print(tokens)
        # get account id
        if tokens[0] == "FUND":
            account = "Unit" if len(tokens) == 7 else tokens[7]
        if "Plan" in tokens:
            # GLOBAL COMMENTS strings mess with this part
            index = next(i for i, s in enumerate(tokens) if "Plan" in s)
            plan = int(tokens[index + 1])
            comments = " ".join(tokens[(index + 2) : :])
        # check for truncated data on the next line
        elif plan > 0 and plan != 1:
            comments += f" {' '.join(tokens)}"
        else:
            continue
        # print(plan, comments)
        rows.append({"Plan": plan, "Comments": comments})
    return rows, account


def parse_pdf(pdf_file, output_file):
    # extract text from "STATUS OF FUNDS" pages
    with pdfplumber.open(pdf_file) as pdf:
        sub_accounts = {}
        for page in pdf.pages:
            lines = build_lines(page)
            page_text = "\n".join(lines).upper()
            # get account funds data
            if (
                "STATUS OF FUNDS" in page_text
                and "COMMENTS" not in page_text
                and "EXPIRED" not in page_text
            ):
                data = parse_funds(lines)
                df = pd.DataFrame(data[0])
                account, account_name = data[1], data[2]
                if account_name != "Unit":
                    sub_accounts[account] = account_name
                if (
                    account == 0
                ):  # First sheet is always the Totals, write it to a new file
                    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                        df.to_excel(writer, sheet_name=f"Unit Summary", index=False)
                else:
                    with pd.ExcelWriter(
                        output_file,
                        engine="openpyxl",
                        mode="a",
                    ) as writer:
                        df.to_excel(
                            writer, sheet_name=f"{account} Summary", index=False
                        )
            # Get the comments data
            if "COMMENTS" in page_text:
                data = parse_comments(lines)
                # skip empty pages
                if not data[0]:
                    continue
                df = pd.DataFrame(data[0])
                account = data[1]
                with pd.ExcelWriter(
                    output_file,
                    engine="openpyxl",
                    mode="a",
                ) as writer:
                    df.to_excel(writer, sheet_name=f"{account} Comments", index=False)
    # handle Excel file - optional
    xlsx_file = pdf_file.replace(".pdf", ".xlsx")
    if os.path.exists(xlsx_file):
        parse_xlsx(xlsx_file, output_file, sub_accounts)
    else:
        if not cmdline:
            messagebox.showwarning(
                "Warning", f"Excel file: {os.path.basename(xlsx_file)} does not exist."
            )
            messagebox.showwarning(
                "Warning", "The output file will only include account summaries."
            )


def parse_xlsx(xlsx_file, output_file, sub_accounts):
    df = pd.read_excel(xlsx_file)
    for account in sub_accounts.keys():
        if account == 0:
            continue
        sub_df = df[df["Detail Sub Account"] == int(account)].copy()
        # sort data by Object Code
        sub_df = sub_df.sort_values(
            by="Object Class", key=lambda x: x.astype(str).str[:3].astype(int)
        )
        with pd.ExcelWriter(
            output_file,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            sub_df.to_excel(
                writer, sheet_name=f"{account} {sub_accounts.get(account)}", index=False
            )


def format_workbook(
    filename,
    currency_columns=None,
    column_padding=3,
    currency_format="$#,##0.00",
):
    if currency_columns is None:
        currency_columns = []
    wb = openpyxl.load_workbook(filename)
    green_fill = openpyxl.styles.PatternFill(
        fill_type="solid", start_color="C6EFCE", end_color="C6EFCE"
    )
    yellow_fill = openpyxl.styles.PatternFill(
        fill_type="solid", start_color="FFEB9C", end_color="FFEB9C"
    )
    red_fill = openpyxl.styles.PatternFill(
        fill_type="solid", start_color="FFC7CE", end_color="FFC7CE"
    )
    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            header = ws.cell(row=1, column=col).value
            if header is None:
                header = ""
            header = str(header)
            column_letter = openpyxl.utils.get_column_letter(col)
            # Fit width to header
            ws.column_dimensions[column_letter].width = len(header) + column_padding
            if header not in currency_columns:
                continue
            cell_range = f"{column_letter}2:{column_letter}{max_row}"
            # Apply currency number format
            for row in range(2, max_row + 1):
                ws.cell(row=row, column=col).number_format = currency_format
            # add a totals row
            total_row = max_row + 1
            ws[f"A{total_row}"] = "Total"
            ws[f"A{total_row}"].font = openpyxl.styles.Font(bold=True)
            # SUM formula
            total_cell = ws.cell(row=total_row, column=col)
            total_cell.value = f"=SUM({column_letter}2:{column_letter}{max_row})"
            total_cell.number_format = currency_format
            total_cell.font = openpyxl.styles.Font(bold=True)
            # simple conditional formatting
            if "Summary" not in ws.title:
                continue
            # status of funds
            ws.conditional_formatting.add(
                cell_range,
                openpyxl.formatting.rule.CellIsRule(
                    operator="lessThan",
                    formula=["0"],
                    fill=red_fill,
                ),
            )
            ws.conditional_formatting.add(
                cell_range,
                openpyxl.formatting.rule.CellIsRule(
                    operator="equal",
                    formula=["0"],
                    fill=yellow_fill,
                ),
            )
            ws.conditional_formatting.add(
                cell_range,
                openpyxl.formatting.rule.CellIsRule(
                    operator="greaterThan",
                    formula=["0"],
                    fill=green_fill,
                ),
            )
    wb.save(filename)


def main():
    os.chdir(os.path.dirname(__file__))
    # handle PDF file - required
    if len(sys.argv) == 2:
        pdf_file = sys.argv[1]
        cmdline = True
    else:
        cmdline = False
        root = Tk()
        root.withdraw()
        pdf_file = filedialog.askopenfilename(
            initialdir=".",
            title="Select Status of Funds PDF",
            filetypes=[("PDF Files", "*.pdf")],
        )
        if not pdf_file:
            messagebox.showerror("Error", "No PDF file selected.")
            return
    if not os.path.exists(pdf_file):
        print(f"{pdf_file} does not exist.")
        return
    output_file = pdf_file.replace(".pdf", "-PARSED.xlsx")
    parse_pdf(pdf_file, output_file)
    # format the new workbook
    format_workbook(
        output_file,
        currency_columns=[
            "FinancialPlan",
            "Reconciled",
            "Unreconciled",
            "TotalObligations",
            "BalanceAvailable",
            "Unreconciled Amt",
            "Reconciled Amt",
        ],
    )
    # success?
    if os.path.exists(output_file):
        try:
            with pd.ExcelFile(output_file) as xls:
                if cmdline:
                    print("Success!")
                else:
                    messagebox.showinfo(
                        "Success",
                        f"File {os.path.basename(output_file)} saved successfully!",
                    )
        except Exception as e:
            if not cmdline:
                messagebox.showerror(
                    "Error", f"Output file {output_file} is corrupted."
                )
    else:
        if not cmdline:
            messagebox.showerror(
                "Error", f"Failed to create output file {output_file}."
            )


if __name__ == "__main__":
    main()
